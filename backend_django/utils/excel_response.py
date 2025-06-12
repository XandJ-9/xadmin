
# -*- coding: utf-8 -*-
from __future__ import absolute_import, unicode_literals

from urllib.parse import quote
from django.http import HttpResponse
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.writer.excel import save_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
import csv
import django
import six

if django.VERSION >= (1, 9):
    from django.db.models.query import QuerySet
else:
    from django.db.models.query import QuerySet, ValuesQuerySet


ROW_LIMIT = 1048576
COL_LIMIT = 16384

DEFAULT_FONT = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none',)

class ExcelResponse(HttpResponse):
    def __init__(self, *args, **kwargs):
        filename = kwargs.pop('filename', 'export.xlsx')
        super().__init__(*args, **kwargs)
        self.headers['Content-Type'] = 'application/msexcel'
        self.headers['content-disposition'] =  f'attachment;filename={quote(str(f"{filename}"))}'
        # # cross-origin跨域请求需要设置Access-Control-Expose-Headers响应信息
        self.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'


class WrapperExcelResponse(HttpResponse):
    """
    This class provides an HTTP Response in the form of an Excel spreadsheet, or CSV file.
    """

    def __init__(self, data, output_filename='excel_data', worksheet_name=None, force_csv=False, header_font=None,
                 data_font=None, guess_types=True, *args, **kwargs):
        # We do not initialize this with streaming_content, as that gets generated when needed
        self.output_filename = output_filename
        self.worksheet_name = worksheet_name or 'Sheet 1'
        self.header_font = header_font or DEFAULT_FONT
        self.data_font = data_font or DEFAULT_FONT
        self.force_csv = force_csv
        self.guess_types = guess_types
        super(ExcelResponse, self).__init__(data, *args, **kwargs)

    @property
    def content(self):
        return b''.join(self._container)

    @content.setter
    def content(self, value):
        workbook = None
        if not bool(value) or not len(value):  # Short-circuit to protect against empty querysets/empty lists/None, etc
            self._container = []
            return
        elif isinstance(value, list):
            workbook = self._serialize_list(value)
        elif isinstance(value, QuerySet):
            workbook = self._serialize_queryset(value)
        if django.VERSION < (1, 9):
            if isinstance(value, ValuesQuerySet):
                workbook = self._serialize_values_queryset(value)
        if workbook is None:
            raise ValueError('ExcelResponse accepts the following data types: list, dict, QuerySet, ValuesQuerySet')

        if self.force_csv:
            self['Content-Type'] = 'text/csv; charset=utf8'
            self['Content-Disposition'] = 'attachment;filename="{}.csv"'.format(self.output_filename)
            workbook.seek(0)
            workbook = self.make_bytes(workbook.getvalue())
        else:
            self['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            self['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(self.output_filename)
            # workbook = Workbook(write_only=True)
        self._container = [self.make_bytes(workbook)]

    def _serialize_list(self, data):
        if isinstance(data[0], dict):  # If we're dealing with a list of dictionaries, generate the headers
            headers = [key for key in data[0]]
        else:
            headers = data[0]
        if len(data) > ROW_LIMIT or len(headers) > COL_LIMIT or self.force_csv:
            self.force_csv = True
            workbook = six.StringIO()
            csvwriter = csv.writer(workbook, dialect='excel')
            append = getattr(csvwriter, 'writerow')
            write_header = append
        else:
            workbook = Workbook(write_only=True)
            workbook.guess_types = self.guess_types
            worksheet = workbook.create_sheet(title=self.worksheet_name)

            # Define custom functions for appending so that we can handle any formatting
            def append(data):
                return self._append_excel_row(worksheet, data, header=False)

            def write_header(data):
                return self._append_excel_row(worksheet, data, header=True)

        if isinstance(data[0], dict):
            append(headers)
        for index, row in enumerate(data):
            if isinstance(row, dict):
                write_header([row.get(col, None) for col in headers])
            else:
                if index > 0:
                    append(row)
                else:
                    write_header(row)
        return workbook

    def _serialize_queryset(self, data):
        if isinstance(data[0], dict):  # .values() returns a list of dicts
            return self._serialize_list(list(data))
        else:
            return self._serialize_list(list(data.values()))

    def _serialize_values_queryset(self, data):
        return self._serialize_list(list(data))

    def _append_excel_row(self, worksheet, data, header=False):
        if header:
            font = self.header_font
        else:
            font = self.data_font

        row = []
        print(f'append_excel_row: {data}')
        for item in data:
            cell = WriteOnlyCell(worksheet, item)
            cell.font = font
            row.append(cell)

        worksheet.append(row)