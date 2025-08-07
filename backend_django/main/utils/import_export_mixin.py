# -*- coding: utf-8 -*-
import datetime
from urllib.parse import quote

from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.table import Table, TableStyleInfo
from rest_framework.decorators import action
from rest_framework.request import Request

from main.utils.util_response import DetailResponse, ExcelResponse, ErrorResponse
from main.utils.util_request import get_verbose_name
from main.utils.models import get_field_verbose_name

from openpyxl import load_workbook

class ImportViewSetMixin:
    import_serializer_class = None
    
    def get_import_serializer_class(self):
        return self.import_serializer_class or self.get_serializer_class()

    def get_import_serializer(self, *args, **kwargs):
        serializer_class = self.get_import_serializer_class()
        return serializer_class(*args, **kwargs)

    @action(methods=['post'], detail=False)
    def import_data(self, request: Request, *args, **kwargs):
        pass

class ExportViewSetMixin:
    export_serializer_class = None

    def get_export_serializer_class(self):
        return self.export_serializer_class or self.get_serializer_class()
    
    def get_export_serializer(self, *args, **kwargs):
        serializer_class = self.get_export_serializer_class()
        return serializer_class(*args, **kwargs)
    @action(methods=['get'], detail=False)
    def export_data(self, request: Request, *args, **kwargs):
        pass

class ExcelImportExportMixin(ImportViewSetMixin, ExportViewSetMixin):
    excel_column_width = 50

    @action(methods=['post'], detail=False, url_name='export_data', url_path='export')
    def export_data(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        assert self.export_field_label, "'%s' 请配置对应的导出模板字段。" % self.__class__.__name__
        serializer = self.get_export_serializer(queryset, many = True, request = request)
        wb = self.make_export_excel(data = serializer.data, *args, **kwargs)
        filename = self.request.query_params.get('file_name', 'export')
        response = ExcelResponse(filename=filename, workbook = wb)
        return response
    
    @action(methods =['post'], detail=False, url_name='import_data', url_path='import')
    def import_data(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')
        update_support = request.query_params.get('updateSupport', False)
        if not file_obj:
            return DetailResponse(status=400, msg='请上传文件')
        data_list = self.parse_import_excel(file_obj)
        if data_list is None:
            return ErrorResponse(status=400, msg='请上传正确的文件')
        for data in data_list:
            if data.get('id', None):
                if update_support:
                    instance = self.get_queryset().filter(id=data.get('id')).first()
                    serializer = self.get_import_serializer(instance = instance,data=data, request=request)
                    if serializer.is_valid(raise_exception=True):
                        serializer.save()
                else:
                    continue
            else:
                serializer = self.get_import_serializer(data=data, request=request)
                if serializer.is_valid(raise_exception=False):
                    serializer.save()
                else:
                    return ErrorResponse(status=400, msg=f'数据验证失败:{data}')
        return DetailResponse(msg='导入成功')

    def make_export_excel(self, data: list = None, *args, **kwargs):
        raise NotImplementedError('子类必须实现make_export_excel方法')

    def parse_import_excel(self, file, *args, **kwargs):
        raise NotImplementedError('子类必须实现parse_import_excel方法')

    def is_number(self,num):
        try:
            float(num)
            return True
        except ValueError:
            pass

        try:
            import unicodedata
            unicodedata.numeric(num)
            return True
        except (TypeError, ValueError):
            pass
        return False

    def get_string_len(self, string):
        """
        获取字符串最大长度
        :param string:
        :return:
        """
        length = 4
        if string is None:
            return length
        if self.is_number(string):
            return length
        for char in string:
            length += 2.1 if ord(char) > 256 else 1
        return round(length, 1) if length <= self.excel_column_width else self.excel_column_width

class ModelImportExportMixin(ExcelImportExportMixin):
    """
    自定义导入模板、导入功能
    """

    # 导入字段
    import_field_dict = {}

    # 导出字段
    export_field_label = {}
    
    def get_export_column_label(self):
        query_set = self.get_queryset()
        model = query_set.model
        result = {}
        for field in model._meta.fields:
            result[field.name] = get_field_verbose_name(model, field.name)
        print(result)

    @action(methods=['post'],detail=False, url_path='template')
    def import_excel_template(self,request):
        queryset = self.filter_queryset(self.get_queryset())
        assert self.import_field_dict, "'%s' 请配置对应的导入模板字段。" % self.__class__.__name__
        serializer = self.get_import_serializer_class()
        data = serializer(queryset, many=True, request=request).data
        wb = self.make_import_template(data)
        response = ExcelResponse(filename=f"导入{get_verbose_name(queryset)}数据模板.xlsx",workbook=wb)
        return response
    
    def make_export_excel(self, data = None, *args, **kwargs):
        self.get_export_column_label()
        wb = Workbook()
        ws = wb.active
        header_data = ["序号", *self.export_field_label.values()]
        hidden_header = ["#", *self.export_field_label.keys()]
        df_len_max = [self.get_string_len(ele) for ele in header_data]
        row = get_column_letter(len(self.export_field_label) + 1)
        column = 1
        ws.append(header_data)
        for index, results in enumerate(data):
            results_list = []
            for h_index, h_item in enumerate(hidden_header[1:]):
                if h_item in results.keys():
                    val = results.get(h_item, "")
                    results_list.append(val)
                    # 　计算最大列宽度
                    result_column_width = self.get_string_len(val)
                    if h_index != 0 and result_column_width > df_len_max[h_index]:
                        df_len_max[h_index] = result_column_width
                else:
                    results_list.append("")
            ws.append([index + 1, *results_list])
            column += 1
        # 　更新列宽
        for index, width in enumerate(df_len_max):
            ws.column_dimensions[get_column_letter(index + 1)].width = width
        tab = Table(displayName="Table", ref=f"A1:{row}{column}")  # 名称管理器
        style = TableStyleInfo(
            name="TableStyleLight11",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        return wb
    def make_import_template(self, data, *args, **kwargs):
        wb = Workbook()
        ws1 = wb.create_sheet("data", 1)
        ws1.sheet_state = "hidden"
        ws = wb.active
        header_data = ["序号","更新主键(勿改)"]
        hidden_header = ["#","id"]
        #----设置选项----
        validation_data_dict = {}
        for index, item in enumerate(self.import_field_dict.items()):
            items = list(item)
            key = items[0]
            value = items[1]
            # key：  记录的字段名
            hidden_header.append(key)
            if isinstance(value, dict):
                header_data.append(value.get("title"))
                # hidden_header.append(value.get('display'))
                choices = value.get("choices", {})
                if choices.get("data"):
                    data_list = []
                    data_list.extend(choices.get("data").keys())
                    validation_data_dict[value.get("title")] = data_list
                elif choices.get("queryset") and choices.get("values_name"):
                    data_list = choices.get("queryset").values_list(choices.get("values_name"), flat=True)
                    validation_data_dict[value.get("title")] = list(data_list)
                else:
                    continue
                column_letter = get_column_letter(len(validation_data_dict))
                dv = DataValidation(
                    type="list",
                    formula1=f"{quote_sheetname('data')}!${column_letter}$2:${column_letter}${len(validation_data_dict[value.get('title')]) + 1}",
                    allow_blank=True,
                )
                ws.add_data_validation(dv)
                dv.add(f"{get_column_letter(index + 3)}2:{get_column_letter(index + 3)}1048576")
            else:
                header_data.append(value)
        # 添加数据列
        ws1.append(list(validation_data_dict.keys()))
        for index, validation_data in enumerate(validation_data_dict.values()):
            for inx, ele in enumerate(validation_data):
                ws1[f"{get_column_letter(index + 1)}{inx + 2}"] = ele
        # --------
        df_len_max = [self.get_string_len(ele) for ele in header_data]
        column = get_column_letter(len(hidden_header))
        row = 1
        ws.append(header_data)
        #  写入已经存在的数据
        for index, results in enumerate(data):
            results_list = []
            for h_index, h_item in enumerate(hidden_header):
                if h_item == '#':
                    continue
                val = results.get(h_item,"")
                if isinstance(val, datetime.datetime):
                    results_list.append(val.strftime('%Y-%m-%d %H:%M:%S'))
                else:
                    results_list.append(str(val))
                # 计算最大列宽度
                if isinstance(val,str):
                    result_column_width = self.get_string_len(val)
                    if h_index != 0 and result_column_width > df_len_max[h_index]:
                        df_len_max[h_index] = result_column_width
            ws.append([str(index+1),*results_list])
            row += 1
        # 　更新列宽
        for index, width in enumerate(df_len_max):
            ws.column_dimensions[get_column_letter(index + 1)].width = width
        tab = Table(displayName="Table", ref=f"A1:{column}{row}")  # 名称管理器
        style = TableStyleInfo(
            name="TableStyleLight11",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        return wb
    def parse_import_excel(self, file, *args, **kwargs):
        wb = load_workbook(file)
        ws = wb.active
        ## 获取字段名
        header_data = ["序号","更新主键(勿改)"]
        hidden_header = ["#","id"]
        for index, item in enumerate(self.import_field_dict.items()):
            key = item[0]
            value = item[1]
            hidden_header.append(key)
            if isinstance(value, dict):
                header_data.append(value.get("title"))
            else:
                header_data.append(value)
        data_list = []
        for row in ws.iter_rows(min_row=2):
            data = {}
            for index, item in enumerate(row):
                if index == 0:
                    continue
                if item.value is None:
                    continue
                data[hidden_header[index]] = item.value
            data_list.append(data)
        # print(f'获取到数据行数：{data_list}')
        return data_list

