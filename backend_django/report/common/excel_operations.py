from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font,DEFAULT_FONT, PatternFill, Border, Side, Alignment
import os
from report.serializers import *
from report.models import *

DefaultStyle={
    "font" : Font(name='Calibri',
                    size=11,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000'),
    'fill' : PatternFill(patternType='solid',fgColor="99CCFF"),
    'border' : Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin',color='FF000000'),
                    top=Side(border_style='thin',color='FF000000'),
                    bottom=Side(border_style='thin',color='FF000000')
                )
}

def set_area_border(ws,start_row,end_row,start_col,end_col):
    for row in range(start_row,end_row+1):
        for col in range(start_col,end_col+1):
            ws.cell(row=row,column=col).border = DefaultStyle['border']

def generate_interface_workbook(interface, fields):

    data = InterfaceInfoSerializer(interface).data
    wb = Workbook()
    ws = wb.active
    ws.title = "report"
    template_info = [
        {'position':'A1','name':'模块名称'},
        {'position':'C1','name':'报表名称'},
        {'position':'E1','name':'报表平台'},
        {'position':'A2','name':'接口名称'},
        {'position':'C2','name':'接口代码'},
        {'position':'E2','name':'日期选项'},
        {'position':'G2','name':'二级表头'},
        {'position':'I2','name':'需要登录'},
        {'position':'K2','name':'告警方式'},
        {'position':'A3','name':'数据库类型'},
        {'position':'C3','name':'数据库名称'},
        {'position':'E3','name':'接口sql'},
        {'position':'A4','name':'是否分页'},
        {'position':'C4','name':'是否合计'},
        {'position':'E4','name':'合计sql'},

    ]
    for info in template_info:
        cell = ws[info['position']]
        cell.value = info['name']
        cell.font = DefaultStyle['font']
        cell.fill = DefaultStyle['fill']
        cell.border = DefaultStyle['border']
    ws['B1']=data['report_info']['module_info']['name']
    ws['D1']=data['report_info']['name']
    ws['F1']=data['report_info']['module_info']['platform_info']['name']
    ws['B2']=data['interface_name']
    ws['D2']=data['interface_code']
    ws['F2']=data['is_date_option']
    ws['H2']=data['is_second_table']
    ws['J2']=data['is_login_visit']
    ws['L2']=data['alarm_type']
    ws['B3']=data['interface_db_type']
    ws['D3']=data['interface_db_name']
    ws['F3']=data['interface_sql']
    ws['B4']=data['is_paging']
    ws['D4']=data['is_total']
    ws['F4']=data['total_sql']
    column_headers = [
        {'label':'序号', 'code':'interface_para_position'},
        {'label':'参数名称','code':'interface_para_name'},
        {'label':'参数代码','code':'interface_para_code'},
        {'label':'参数类型','code':'interface_para_type'},
        {'label':'数据类型','code':'interface_data_type'},
        {'label':'是否展示','code':'interface_show_flag'},
        {'label':'是否导出','code':'interface_export_flag'},
        {'label':'参数接口代码','code':'interface_para_interface_code'},
        {'label':'参数默认值','code':'interface_para_default'},
        {'label':'级联参数','code':'interface_cascade_para'},
        {'label':'父表头名称','code':'interface_parent_name'},
        {'label':'父表头位置','code':'interface_parent_position'},
        {'label':'是否合并行','code':'interface_para_rowspan'},
        {'label':'是否显示备注','code':'interface_show_desc'},
        {'label':'参数描述','code':'interface_para_desc'}
        ]
    for index,column_info in enumerate(column_headers):
        cell = ws.cell(row=5,column=index+1)
        cell.font = DefaultStyle['font']
        cell.fill = PatternFill(patternType='solid', fgColor='C0C0C0')
        cell.value = column_info.get('label')
    if fields:
        fields_data = InterfaceFieldSerializer(instance=fields,many=True).data
        for row_index,field in enumerate(fields_data):
            for col_index,column_info in enumerate(column_headers):
                ws.cell(row=row_index+6,column=1+col_index).value = field[column_info['code']]
    set_area_border(ws,ws.min_row,ws.max_row,ws.min_column,ws.max_column)
    ws['P1']='report'
    return wb


def handle_interface_import(full_filepath, user = None):
    if not os.path.exists(full_filepath):
        raise Exception(f'{full_filepath} not exists')
    filename = os.path.basename(full_filepath)
    wb = load_workbook(full_filepath)
    ws = wb['report']
    platform_info = {
        'name':ws['F1'].value,
        'desc':ws['G1'].value
    }
    platform, created = PlatformInfo.objects.get_or_create(name=platform_info['name'],
                                                           defaults={
                                                               'desc': platform_info['desc'],
                                                               'creator': user
                                                           })
    if not created:
        # 数据表中已存在，则更新
        platform.desc = platform_info['desc']
        platform.updator = user
        platform.save()
    
    module_info = {
        'name':ws['B1'].value,
        'platform': platform
    }
    module, created = ModuleInfo.objects.get_or_create(name=module_info['name'],
                                                       defaults={
                                                           'platform':module_info['platform'], 
                                                           'creator':user
                                                        })
    if not created:
        module.platform = module_info['platform']
        module.updator = user
        module.save()
    
    report_info = {
        'name':ws['D1'].value,
        'module':module
    }
    report, created = ReportInfo.objects.get_or_create(name=report_info['name'],
                                                       defaults ={
                                                           'module':report_info['module'],
                                                           'creator':user
                                                       })
    if not created:
        report.module = report_info['module']
        report.updator = user
        report.save()
    
    # 下述方法使用模型序列化类来保存数据
    interface_info = {
        'interface_name':ws['B2'].value,
        'interface_code':ws['D2'].value,
        'is_date_option':ws['F2'].value,
        'is_second_table':ws['H2'].value,
        'is_login_visit':ws['J2'].value,
        'alarm_type':ws['L2'].value,
        'interface_db_type':ws['B3'].value,
        'interface_db_name':ws['D3'].value,
        'interface_sql':ws['F3'].value,
        'is_paging':ws['B4'].value,
        'is_total':ws['D4'].value,
        'total_sql':ws['F4'].value,
        'report':report.id,
        'creator':user.id,
        'updator':user.id
    }
    interface = InterfaceInfo.objects.filter(interface_code = interface_info['interface_code'])
    if interface:
        ser = InterfaceInfoSerializer(instance = interface[0],data=interface_info)
    else:
        ser = InterfaceInfoSerializer(data=interface_info)
    # 在序列化模型对象时，会将选项字段转换
    ser.is_valid(raise_exception=True)
    ser.save()

    # 导入字段信息
    field_row_start = 6 
    column_headers = [
        {'label':'序号', 'code':'interface_para_position'},
        {'label':'参数名称','code':'interface_para_name'},
        {'label':'参数代码','code':'interface_para_code'},
        {'label':'参数类型','code':'interface_para_type'},
        {'label':'数据类型','code':'interface_data_type'},
        {'label':'是否展示','code':'interface_show_flag'},
        {'label':'是否导出','code':'interface_export_flag'},
        {'label':'参数接口代码','code':'interface_para_interface_code'},
        {'label':'参数默认值','code':'interface_para_default'},
        {'label':'级联参数','code':'interface_cascade_para'},
        {'label':'父表头名称','code':'interface_parent_name'},
        {'label':'父表头位置','code':'interface_parent_position'},
        {'label':'是否合并行','code':'interface_para_rowspan'},
        {'label':'是否显示备注','code':'interface_show_desc'},
        {'label':'参数描述','code':'interface_para_desc'}
        ]
    # 读取excel中的字段信息
    interface_fields = []
    for row_idx in range(field_row_start,ws.max_row+1):
        field_info = {'interface':ser.data['id'],'creator':user.id,'updator':user.id}
        for col_idx in range(min(len(column_headers),ws.max_column)):
            field_info[column_headers[col_idx]['code']] = ws.cell(row=row_idx,column=col_idx+1).value
        interface_fields.append(field_info)
    
    existed_fields = InterfaceField.objects.filter(interface_id = ser.data['id'])

    field_ser = InterfaceFieldSerializer(data=interface_fields,many=True)
    field_ser.is_valid(raise_exception=True)
    new_fields = field_ser.save()

    need_delete_fields = existed_fields.exclude(id__in=[field.id for field in new_fields])
    need_delete_fields.delete()
