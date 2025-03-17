from rest_framework.viewsets import ViewSet
from rest_framework.decorators import action

from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Font,DEFAULT_FONT, PatternFill, Border, Side, Alignment
from urllib.parse import quote

from report.models import Interface, InterfaceField, ReportInfo
from report.utils.util_response import SuccessResponse, DetailResponse
from report.serializers import InterfaceSerializer, InterfaceFieldSerializer

DefaultStyle={
    "font" : Font(name='Calibri',
                    size=11,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000'),
    'fill' : PatternFill(patternType='solid',fgColor="99CCFF"),
    'border' : Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin',color='FF000000'),
                    top=Side(border_style='thin',color='FF000000'),
                    bottom=Side(border_style='thin',color='FF000000')
                )
}

def set_area_border(ws,start_row,end_row,start_col,end_col):
    for row in range(start_row,end_row+1):
        for col in range(start_col,end_col+1):
            ws.cell(row=row,column=col).border = DefaultStyle['border']

class ImportViewSet(ViewSet):
    pass


def generate_interface_workbook(interface_id):
    instance = Interface.objects.get(id = interface_id)
    
    if not instance:
        return DetailResponse(code=400,msg='接口不存在')
    data = InterfaceSerializer(instance).data
    wb = Workbook()
    ws = wb.active
    ws.title = "report"
    template_info = [
        {'position':'A1','name':'模块名称'},
        {'position':'C1','name':'报表名称'},
        {'position':'E1','name':'报表平台'},
        {'position':'A2','name':'接口名称'},
        {'position':'C2','name':'接口代码'},
        {'position':'E2','name':'日期选项'},
        {'position':'G2','name':'二级表头'},
        {'position':'I2','name':'需要登录'},
        {'position':'K2','name':'告警方式'},
        {'position':'A3','name':'数据库类型'},
        {'position':'C3','name':'数据库名称'},
        {'position':'E3','name':'接口sql'},
        {'position':'A4','name':'是否分页'},
        {'position':'C4','name':'是否合计'},
        {'position':'E4','name':'合计sql'},

    ]
    for info in template_info:
        cell = ws[info['position']]
        cell.value = info['name']
        cell.font = DefaultStyle['font']
        cell.fill = DefaultStyle['fill']
        cell.border = DefaultStyle['border']
    ws['B1']=data['module']
    ws['D1']=data['report']
    ws['F1']=data['platform']
    ws['B2']=data['interface_name']
    ws['D2']=data['interface_code']
    ws['F2']=data['is_date_option']
    ws['H2']=data['is_second_table']
    ws['J2']=data['is_login_visit']
    ws['L2']=data['alarm_type']
    ws['B3']=data['interface_db_type']
    ws['D3']=data['interface_db_name']
    ws['F3']=data['interface_sql']
    ws['B4']=data['is_paging']
    ws['D4']=data['is_total']
    ws['F4']=data['total_sql']
    column_headers = [
        {'label':'序号', 'code':'interface_para_position'},
        {'label':'参数名称','code':'interface_para_name'},
        {'label':'参数代码','code':'interface_para_code'},
        {'label':'参数类型','code':'interface_para_type'},
        {'label':'数据类型','code':'interface_data_type'},
        {'label':'是否展示','code':'interface_show_flag'},
        {'label':'是否导出','code':'interface_export_flag'},
        {'label':'参数接口代码','code':'interface_para_interface_code'},
        {'label':'参数默认值','code':'interface_para_default'},
        {'label':'级联参数','code':'interface_cascade_para'},
        {'label':'父表头名称','code':'interface_parent_name'},
        {'label':'父表头位置','code':'interface_parent_position'},
        {'label':'是否合并行','code':'interface_para_rowspan'},
        {'label':'是否显示备注','code':'interface_show_desc'},
        {'label':'参数描述','code':'interface_para_desc'}
        ]
    for index,column_info in enumerate(column_headers):
        cell = ws.cell(row=5,column=index+1)
        cell.font = DefaultStyle['font']
        cell.fill = PatternFill(patternType='solid', fgColor='C0C0C0')
        cell.value = column_info.get('label')
    interfaceFields = InterfaceField.objects.filter(interface_id=interface_id)
    fields_data = InterfaceFieldSerializer(instance=interfaceFields,many=True).data
    for row_index,field in enumerate(fields_data):
        for col_index,column_info in enumerate(column_headers):
            ws.cell(row=row_index+6,column=1+col_index).value = field[column_info['code']]
    set_area_border(ws,ws.min_row,ws.max_row,ws.min_column,ws.max_column)
    ws['P1']='report'
    return wb

class ExportViewSet(ViewSet):
    # authentication_classes =[]
    @action(methods=['post'], detail=False, url_path='interface')
    def export_interface(self,request):
        interface_id = request.query_params.get('interface_id')
        instance = Interface.objects.get(id = interface_id)
        
        if not instance:
            return DetailResponse(code=400,msg='接口不存在')
        response = HttpResponse(content_type="application/msexcel")
        ## cross-origin跨域请求需要设置Access-Control-Expose-Headers响应信息
        response["Access-Control-Expose-Headers"] = f"Content-Disposition"
        response["content-disposition"] = f'attachment;filename={quote(str(f"{instance.report}-{instance.interface_name}.xlsx"))}'
        wb = generate_interface_workbook(interface_id)
        wb.save(response)
        return response
    
    def export_multi_interface(self, request):
        interface_id_list = request.data.get('interface_id_list')
        for interface_id in interface_id_list:
            wb=generate_interface_workbook(interface_id=interface_id)
            